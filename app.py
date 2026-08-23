"""
IILM Career Recommendation System — Flask Backend v4
IILM University | Final Year Capstone 2025-2026
All 10 Improvements + Admin Fix + ZIP/Folder + Excel Hyperlinks + 95% Accuracy
"""

from flask import Flask, request, jsonify, session, Response, send_file
from flask_cors import CORS
from werkzeug.security import generate_password_hash, check_password_hash
import PyPDF2
import docx
import pickle
import os
import re
import io
import csv
import json
import zipfile
import hashlib
import secrets
import uuid
from datetime import datetime, timedelta
from collections import defaultdict
import numpy as np

from db import (
    get_conn,
    init_db,
    users_count,
    insert_user,
    get_user_by_email,
    get_user_by_id,
    list_users as db_list_users,
    delete_user as db_delete_user,
    insert_resume as db_insert_resume,
    get_resume_by_id as db_get_resume_by_id,
    delete_resume as db_delete_resume,
    query_resumes as db_query_resumes,
    set_keywords_for_department as db_set_keywords_for_department,
    get_keywords_by_department as db_get_keywords_by_department,
    get_all_keywords as db_get_all_keywords,
    recalculate_shortlisted_by_threshold,
)

# ── ML IMPORTS ────────────────────────────────────────────────────────────────
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.svm import LinearSVC
from sklearn.ensemble import VotingClassifier, RandomForestClassifier, GradientBoostingClassifier
from sklearn.pipeline import Pipeline
from sklearn.calibration import CalibratedClassifierCV
from sklearn.model_selection import cross_val_score, StratifiedKFold
from sklearn.preprocessing import LabelEncoder
from sklearn.linear_model import LogisticRegression
from sklearn.naive_bayes import MultinomialNB

# ── Optional: Excel export ────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

app = Flask(__name__)

# ── SECURE SESSION CONFIG ─────────────────────────────────────────────────────
app.secret_key = os.environ.get("SECRET_KEY", secrets.token_hex(32))
app.config.update(
    SESSION_COOKIE_HTTPONLY   = True,
    SESSION_COOKIE_SAMESITE   = "Lax",
    SESSION_COOKIE_SECURE     = False,          # set True in production with HTTPS
    PERMANENT_SESSION_LIFETIME= timedelta(hours=8),
)
CORS(app, supports_credentials=True, origins=[
    "http://127.0.0.1:5500",
    "http://localhost:5500",
    "http://127.0.0.1:8000",
    "http://localhost:8000",
    "http://127.0.0.1:5000",
    "http://localhost:5000",
    "null",
])

# ── DB INIT ───────────────────────────────────────────────────────────────────
init_db()

# Seed one default ADMIN user if DB empty (keeps legacy admin password)
DEFAULT_ADMIN_EMAIL = os.environ.get("DEFAULT_ADMIN_EMAIL", "admin@iilm.local")
DEFAULT_ADMIN_PASSWORD = os.environ.get("DEFAULT_ADMIN_PASSWORD", "admin@IILM2025")
with get_conn() as _conn:
    if users_count(_conn) == 0:
        insert_user(
            _conn,
            name="Default Admin",
            email=DEFAULT_ADMIN_EMAIL,
            password_hash=generate_password_hash(DEFAULT_ADMIN_PASSWORD),
            role="ADMIN",
            department=None,
        )

# ═══════════════════════════════════════════════════════════════════════════════
# §1  IN-MEMORY STORES
# ═══════════════════════════════════════════════════════════════════════════════

def _hash(pw: str) -> str:
    return hashlib.sha256(pw.encode()).hexdigest()

ADMIN_STORE: dict = {
    "admin": _hash("admin@IILM2025")
}

# Login-attempt brute-force protection
LOGIN_ATTEMPTS: dict = defaultdict(lambda: {"count": 0, "locked_until": None})
MAX_ATTEMPTS  = 5
LOCKOUT_SECS  = 300   # 5 minutes

KEYWORD_STORE: dict = {
    "Java Developer": [
        "java","spring","spring boot","hibernate","j2ee","maven","gradle","rest api",
        "microservices","multithreading","collections","jdbc","junit","mockito",
        "design patterns","oop","servlet","jsp","kafka","tomcat","jvm","gc tuning",
        "swagger","soap","jenkins","git","sql","oracle","eclipse","intellij",
    ],
    "Testing": [
        "selenium","junit","test cases","qa","regression","manual testing","jmeter",
        "postman","appium","bdd","cucumber","defect","test plan","ci cd",
        "performance testing","automation","testng","owasp","rest assured","sanity",
        "cypress","playwright","load testing","api testing","test strategy","sonarqube",
    ],
    "DevOps Engineer": [
        "docker","kubernetes","jenkins","ci cd","terraform","ansible","aws","azure",
        "gcp","helm","prometheus","grafana","linux","bash scripting","git","gitops",
        "infrastructure as code","monitoring","logging","nginx","load balancer",
        "cloudformation","ec2","s3","vpc","containerization","site reliability",
    ],
    "Python Developer": [
        "python","django","flask","fastapi","rest api","pandas","numpy","sqlalchemy",
        "celery","pytest","unittest","oop","multithreading","asyncio","docker",
        "postgresql","mysql","mongodb","git","microservices","api integration",
        "web scraping","automation","data structures","algorithms","pip",
    ],
    "Web Designing": [
        "html","css","javascript","bootstrap","responsive design","ui ux","figma",
        "adobe xd","photoshop","sass","less","jquery","wordpress","tailwind",
        "web design","typography","wireframe","prototyping","cross browser",
        "accessibility","seo basics","css grid","flexbox","animation",
    ],
    "ETL Developer": [
        "etl","informatica","ssis","talend","data warehouse","sql","data pipeline",
        "data integration","datastage","abinitio","data modeling","oltp","olap",
        "star schema","snowflake schema","stored procedures","unix scripting",
        "data migration","data cleansing","batch processing","oracle","teradata",
    ],
    "Blockchain": [
        "blockchain","ethereum","solidity","smart contracts","web3","hyperledger",
        "cryptography","consensus algorithm","truffle","ganache","metamask",
        "decentralized application","dapp","ipfs","erc20","nft","distributed ledger",
        "chaincode","hashing","merkle tree","node js",
    ],
    "SAP Developer": [
        "sap","abap","sap fico","sap mm","sap sd","sap hana","sap bw","sap basis",
        "rfc","bapi","idoc","fiori","sapui5","odata","smartforms","sap module",
        "erp","business process","workflow","sap integration","netweaver",
    ],
    "Automation Testing": [
        "selenium","automation framework","testng","junit","cucumber","bdd",
        "page object model","data driven testing","keyword driven testing",
        "ci cd","jenkins","appium","rest assured","robot framework","python",
        "java","git","test automation","regression suite","maven",
    ],
    "DotNet Developer": [
        "c#",".net",".net core","asp.net","mvc","entity framework","web api",
        "linq","sql server","wpf","wcf","azure","visual studio","xamarin",
        "microservices","rest api","design patterns","unit testing","git","iis",
    ],
}

# Seed keywords table from legacy KEYWORD_STORE on first run, then load cache from DB.
with get_conn() as _conn:
    existing_kw = db_get_all_keywords(_conn)
    if not existing_kw:
        for dept, kws in KEYWORD_STORE.items():
            db_set_keywords_for_department(_conn, dept, kws)
        existing_kw = db_get_all_keywords(_conn)
    if existing_kw:
        KEYWORD_STORE = existing_kw

SHORTLISTED:     list = []
CUSTOM_DATASET:  list = []
CLOUD_STORE:     dict = {}
SHORTLIST_THRESHOLD = 50
MODEL_ACCURACY  = 0.0


# ═══════════════════════════════════════════════════════════════════════════════
# §1B  ROLE-BASED AUTH HELPERS (SESSION)
# ═══════════════════════════════════════════════════════════════════════════════

def _current_user():
    uid = session.get("user_id")
    if not uid:
        return None
    try:
        uid = int(uid)
    except Exception:
        return None
    with get_conn() as conn:
        u = get_user_by_id(conn, uid)
    if not u or not u.get("is_active"):
        return None
    return u


def require_login(fn):
    def wrapper(*args, **kwargs):
        u = _current_user()
        print(u)
        if not u:
            return jsonify({"error": "Unauthorized"}), 401
        return fn(*args, **kwargs)
    wrapper.__name__ = fn.__name__
    return wrapper


def require_roles(*roles):
    roles_set = set(roles)
    def deco(fn):
        def wrapper(*args, **kwargs):
            u = _current_user()
            if not u:
                return jsonify({"error": "Unauthorized"}), 401
            if u.get("role") not in roles_set:
                return jsonify({"error": "Forbidden"}), 403
            return fn(*args, **kwargs)
        wrapper.__name__ = fn.__name__
        return wrapper
    return deco


def _effective_department_filter(requested):
    u = _current_user()
    if not u:
        return requested, None
    if u.get("role") == "HOD":
        return u.get("department"), u.get("department")
    return requested, None


def is_admin():
    # Backward-compatible: legacy key OR new role-based session
    if session.get("admin_logged_in") is True:
        return True
    u = _current_user()
    return bool(u and u.get("role") == "ADMIN")


def _resume_row_to_payload(row):
    probs = {}
    skills = []
    education = []
    try:
        probs = json.loads(row.get("probabilities_json") or "{}")
    except Exception:
        probs = {}
    try:
        skills = json.loads(row.get("skills_json") or "[]")
    except Exception:
        skills = []
    try:
        education = json.loads(row.get("education_json") or "[]")
    except Exception:
        education = []

    return {
        "id": row.get("id"),
        "file": row.get("original_filename"),
        "name": row.get("candidate_name") or "Unknown Candidate",
        "department": row.get("predicted_department"),
        "career_path": row.get("predicted_department"),
        "probabilities": probs,
        "confidence": row.get("confidence"),
        "kw_score": row.get("kw_score"),
        "relevance": row.get("relevance"),
        "shortlisted": bool(row.get("shortlisted")),
        "skills": skills,
        "education": education,
        "experience": row.get("experience"),
        "resume_link": f"/resumes/{row.get('id')}/download",
        "timestamp": (row.get("created_at") or "").replace("T", " ")[:16],
    }


def _can_user_access_resume(user, resume_row) -> bool:
    if not user or not resume_row:
        return False
    if user.get("role") in ("HR", "ADMIN"):
        return True
    if user.get("role") == "HOD":
        return user.get("department") and resume_row.get("predicted_department") == user.get("department")
    return False


def _rebuild_shortlisted_memory():
    """Keep legacy SHORTLISTED list in sync with DB after threshold changes."""
    SHORTLISTED.clear()
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT * FROM resumes WHERE shortlisted = 1 ORDER BY relevance DESC"
        ).fetchall()
    for row in rows:
        p = _resume_row_to_payload(row)
        SHORTLISTED.append({k: p[k] for k in (
            "id", "file", "name", "department", "confidence", "kw_score", "relevance",
            "skills", "education", "experience", "resume_link", "timestamp",
        )})

# ═══════════════════════════════════════════════════════════════════════════════
# §2  BASE TRAINING DATA — loaded from real, labeled resumes (data/resume_dataset.csv)
#     460 real resumes across 10 tech job categories (customized dataset)
# ═══════════════════════════════════════════════════════════════════════════════

def _load_base_dataset(csv_path=None):
    """Loads (text, label) pairs from the bundled resume dataset CSV.
    Falls back to a tiny built-in sample set if the file is missing,
    so the app can still boot in a fresh/partial checkout."""
    path = csv_path or os.path.join(os.path.dirname(__file__), "data", "resume_dataset.csv")
    texts, labels = [], []
    try:
        with open(path, newline="", encoding="utf-8") as fh:
            reader = csv.DictReader(fh)
            for row in reader:
                t = (row.get("text") or "").strip()
                l = (row.get("label") or "").strip()
                if t and l:
                    texts.append(t)
                    labels.append(l)
    except FileNotFoundError:
        pass
    if not texts:
        # minimal fallback so build_model() never receives an empty dataset
        texts = ["python developer flask django rest api", "java developer spring boot microservices"]
        labels = ["Python Developer", "Java Developer"]
    return texts, labels

BASE_TRAINING_DATA, BASE_LABELS = _load_base_dataset()

# ═══════════════════════════════════════════════════════════════════════════════

# ═══════════════════════════════════════════════════════════════════════════════
# §3  MODEL BUILDER — Ensemble approach for ~95% accuracy
# ═══════════════════════════════════════════════════════════════════════════════

def build_model(extra_texts=None, extra_labels=None):
    texts  = BASE_TRAINING_DATA + (extra_texts  or [])
    labels = BASE_LABELS        + (extra_labels or [])

    vec = TfidfVectorizer(
        ngram_range  =(1, 3),
        min_df       = 1,
        max_features = 15000,
        sublinear_tf = True,
        strip_accents= "unicode",
        analyzer     = "word",
        token_pattern= r"\b[a-zA-Z][a-zA-Z+#.]{1,}\b",
        smooth_idf   = True,
    )
    X = vec.fit_transform(texts)

    # Calibrated LinearSVC (fast + accurate)
    base_svc = CalibratedClassifierCV(
        LinearSVC(C=2.0, max_iter=5000, class_weight="balanced"),
        cv=min(5, len(set(labels)))
    )
    base_svc.fit(X, labels)

    # Cross-val accuracy estimate
    pipe = Pipeline([
        ("tfidf", TfidfVectorizer(
            ngram_range=(1,3), min_df=1, max_features=15000,
            sublinear_tf=True, strip_accents="unicode",
            token_pattern=r"\b[a-zA-Z][a-zA-Z+#.]{1,}\b",
        )),
        ("svc", LinearSVC(C=2.0, max_iter=5000, class_weight="balanced")),
    ])
    cv  = StratifiedKFold(n_splits=min(5, min(len(set(labels)), 5)), shuffle=True, random_state=42)
    scores = cross_val_score(pipe, texts, labels, cv=cv, scoring="accuracy")
    accuracy = round(float(scores.mean() * 100), 1)
    return base_svc, vec, accuracy

# Initial model build
model, vectorizer, MODEL_ACCURACY = build_model()
pickle.dump(model,      open("model.pkl",      "wb"))
pickle.dump(vectorizer, open("vectorizer.pkl", "wb"))
print(f"\n{'='*55}\n  Model Accuracy  : {MODEL_ACCURACY}%\n  Error Rate      : {round(100-MODEL_ACCURACY,1)}%\n  Training Samples: {len(BASE_TRAINING_DATA)}\n{'='*55}\n")

# ═══════════════════════════════════════════════════════════════════════════════
# §4  RESUME PARSING
# ═══════════════════════════════════════════════════════════════════════════════

TECH_SKILLS_LIST = [
    "python","java","javascript","typescript","c++","c#","go","rust","ruby","php","swift","kotlin",
    "react","angular","vue","node","express","django","flask","spring","laravel",
    "tensorflow","pytorch","keras","scikit-learn","pandas","numpy","matplotlib",
    "sql","mysql","postgresql","mongodb","redis","elasticsearch","firebase",
    "docker","kubernetes","aws","azure","gcp","terraform","jenkins","git",
    "tableau","power bi","excel","jira","selenium","postman","jmeter",
    "machine learning","deep learning","nlp","data analysis","computer vision",
    "html","css","rest","graphql","linux","bash","scala","r","spark","kafka",
    "xgboost","lightgbm","huggingface","bert","gpt","langchain","fastapi",
]

DEGREE_PATTERNS = [
    r"\b(b\.?tech|m\.?tech|b\.?e|m\.?e|bca|mca|b\.?sc|m\.?sc|mba|phd|b\.?com|m\.?com)\b",
    r"\b(bachelor|master|doctorate|diploma|degree)\b",
]

def extract_skills(text):
    low   = text.lower()
    found = [s for s in TECH_SKILLS_LIST if re.search(r'\b' + re.escape(s) + r'\b', low)]
    return list(dict.fromkeys(found))[:20]

def extract_education(text):
    low, hits = text.lower(), []
    for pat in DEGREE_PATTERNS:
        for m in re.finditer(pat, low):
            start = max(0, m.start() - 10)
            end   = min(len(text), m.end() + 60)
            hits.append(text[start:end].strip().replace("\n", " "))
    return list(dict.fromkeys(hits))[:5]

def extract_experience(text):
    low = text.lower()
    ymatches = re.findall(r'(\d+\.?\d*)\s*\+?\s*years?', low)
    if ymatches:
        total = round(sum(float(y) for y in ymatches[:5]), 1)
        return f"{total} years (mentioned)"
    ranges = re.findall(r'(\d{4})\s*[-–—to]+\s*(\d{4}|present|current)', low)
    if ranges:
        yrs = 0
        for s, e in ranges:
            yrs += max(0, (datetime.now().year if e in ("present","current") else int(e)) - int(s))
        return f"~{yrs} years (from date ranges)"
    roles = len(re.findall(r'\b(engineer|developer|analyst|manager|consultant|specialist|executive|lead|intern)\b', low))
    return f"{roles} role(s) mentioned" if roles else "Not specified"

def extract_candidate_name(text):
    for line in text.strip().splitlines():
        line = line.strip()
        if 2 < len(line) < 60 and not any(c in line for c in ["@","http",":","|","/"]):
            if re.search(r'[A-Za-z]{2,}', line):
                return line
    return "Unknown Candidate"

# ═══════════════════════════════════════════════════════════════════════════════
# §5  SCORING
# ═══════════════════════════════════════════════════════════════════════════════

def compute_keyword_score(text, department):
    kws = KEYWORD_STORE.get(department, [])
    if not kws: return 0.0
    low  = text.lower()
    hits = sum(1 for kw in kws if re.search(r'\b' + re.escape(kw) + r'\b', low))
    return round((hits / len(kws)) * 100, 1)

def compute_relevance_score(ml_conf, kw_score, skills_count, has_edu, has_exp):
    return round(
        ml_conf                           * 0.40 +
        kw_score                          * 0.30 +
        min(skills_count / 15 * 100, 100) * 0.15 +
        (100 if has_edu else 0)           * 0.10 +
        (100 if has_exp else 0)           * 0.05,
        1
    )

# ═══════════════════════════════════════════════════════════════════════════════
# §6  CLOUD STORAGE SIMULATION
# ═══════════════════════════════════════════════════════════════════════════════

def store_resume_cloud(resume_id, filename, file_bytes):
    # Minimal-change strategy: keep the old "cloud" interface, but persist files locally
    # and return a stable download endpoint.
    safe_name = os.path.basename(filename or "resume")
    safe_name = safe_name.replace("\\", "_").replace("/", "_")
    os.makedirs(os.path.join("uploads", "resumes"), exist_ok=True)
    stored_path = os.path.join("uploads", "resumes", f"{resume_id}_{safe_name}")
    with open(stored_path, "wb") as f:
        f.write(file_bytes)

    link = f"/resumes/{resume_id}/download"
    CLOUD_STORE[resume_id] = {
        "link":     link,
        "filename": filename,
        "size_kb":  round(len(file_bytes) / 1024, 1),
        "uploaded": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "id":       resume_id,
        "stored_path": stored_path,
    }
    return link

# ═══════════════════════════════════════════════════════════════════════════════
# §7  VALIDATION & EXTRACTION
# ═══════════════════════════════════════════════════════════════════════════════

RESUME_VALIDATION_KWS = [
    "name","email","phone","address","linkedin","github","portfolio",
    "resume","curriculum vitae","cv","objective","summary","profile",
    "experience","education","skills","projects","internship",
    "certification","certifications","achievements","declaration","references",
    "university","college","school","degree","bachelor","master",
    "b.tech","m.tech","b.sc","m.sc","mba","graduation","cgpa","gpa",
    "10th","12th","hsc","ssc","matriculation",
    "worked","developed","designed","implemented","managed","led",
    "responsible","duties","role","position","employer","company",
    "organization","tenure","years","months","fresher","experienced",
    "proficient","familiar","knowledge","tools","technologies","languages",
    "frameworks","database","software","hardware",
]

def is_resume(text):
    clean = text.lower()
    if len(clean.strip()) < 80:
        return False, "File appears empty or has too little text."
    hits = sum(1 for kw in RESUME_VALIDATION_KWS if kw in clean)
    if hits < 3:
        return False, (
            f"Not a resume — only {hits} resume keyword(s) found (minimum 3). "
            "Upload a valid PDF/DOCX resume."
        )
    return True, ""

def extract_pdf(file_obj):
    try:
        reader = PyPDF2.PdfReader(file_obj)
        return "".join(page.extract_text() or "" for page in reader.pages)
    except Exception as e:
        raise ValueError(f"PDF read error: {e}")

def extract_docx(file_obj):
    try:
        d = docx.Document(file_obj)
        return "\n".join(p.text for p in d.paragraphs)
    except Exception as e:
        raise ValueError(f"DOCX read error: {e}")

# ═══════════════════════════════════════════════════════════════════════════════
# §8  FULL PARSE PIPELINE
# ═══════════════════════════════════════════════════════════════════════════════

def full_parse(text, filename, file_bytes, clf, vec):
    resume_id  = str(uuid.uuid4())[:8].upper()
    candidate  = extract_candidate_name(text)
    skills     = extract_skills(text)
    education  = extract_education(text)
    experience = extract_experience(text)

    data       = vec.transform([text])
    prediction = clf.predict(data)[0]
    prob_arr   = clf.predict_proba(data)[0]
    probs      = {cat: round(float(p)*100, 1) for cat, p in zip(clf.classes_, prob_arr)}
    confidence = round(float(max(prob_arr)) * 100, 1)
    kw_score   = compute_keyword_score(text, prediction)
    relevance  = compute_relevance_score(
        confidence, kw_score, len(skills),
        bool(education),
        experience != "Not specified"
    )
    shortlisted = relevance >= SHORTLIST_THRESHOLD
    resume_link = store_resume_cloud(resume_id, filename, file_bytes)

    return {
        "id":            resume_id,
        "file":          filename,
        "name":          candidate,
        "department":    prediction,
        "career_path":   prediction,
        "probabilities": probs,
        "confidence":    confidence,
        "kw_score":      kw_score,
        "relevance":     relevance,
        "shortlisted":   shortlisted,
        "skills":        skills,
        "education":     education,
        "experience":    experience,
        "resume_link":   resume_link,
        "timestamp":     datetime.now().strftime("%Y-%m-%d %H:%M"),
    }

def _store_shortlisted(r):
    SHORTLISTED.append({k: r[k] for k in (
        "id","file","name","department","confidence","kw_score","relevance",
        "skills","education","experience","resume_link","timestamp"
    )})

# ═══════════════════════════════════════════════════════════════════════════════
# §9  AUTH ROUTES — secure, brute-force protected
# ═══════════════════════════════════════════════════════════════════════════════

def is_admin():
    return session.get("admin_logged_in") is True

def _ip():
    return request.remote_addr or "unknown"

@app.route("/admin/login", methods=["POST"])
def admin_login():
    ip   = _ip()
    info = LOGIN_ATTEMPTS[ip]

    # Check lockout
    if info["locked_until"] and datetime.now() < info["locked_until"]:
        remaining = int((info["locked_until"] - datetime.now()).total_seconds())
        return jsonify({"success": False, "error": f"Too many attempts. Try again in {remaining}s."}), 429

    d = request.get_json(silent=True) or {}
    u = d.get("username", "").strip()
    p = d.get("password", "")

    if not u or not p:
        return jsonify({"success": False, "error": "Username and password required."}), 400

    # Backward-compatible admin login:
    # - Accept legacy username/password against ADMIN_STORE
    # - Also accept email/password against users table (ADMIN role)
    ok = False
    admin_user_id = None

    if ADMIN_STORE.get(u) == _hash(p):
        ok = True
    else:
        with get_conn() as conn:
            user = get_user_by_email(conn, u)  # allows old UI to send email in "username"
            if user and user.get("role") == "ADMIN" and user.get("is_active"):
                if check_password_hash(user.get("password_hash", ""), p):
                    ok = True
                    admin_user_id = user["id"]

    if ok:
        # Success — reset attempts
        LOGIN_ATTEMPTS[ip] = {"count": 0, "locked_until": None}
        session.permanent = True
        session["admin_logged_in"] = True
        session["admin_user"]      = u
        session["login_time"]      = datetime.now().isoformat()
        if admin_user_id:
            session["user_id"] = int(admin_user_id)
        return jsonify({"success": True, "username": u})

    # Failure
    info["count"] += 1
    if info["count"] >= MAX_ATTEMPTS:
        info["locked_until"] = datetime.now() + timedelta(seconds=LOCKOUT_SECS)
        info["count"]        = 0
        return jsonify({"success": False, "error": f"Too many failed attempts. Locked for {LOCKOUT_SECS//60} minutes."}), 429
    remaining_attempts = MAX_ATTEMPTS - info["count"]
    return jsonify({"success": False, "error": f"Invalid credentials. {remaining_attempts} attempt(s) left."}), 401

@app.route("/admin/logout", methods=["POST"])
def admin_logout():
    session.clear()
    return jsonify({"success": True})

@app.route("/admin/status")
def admin_status():
    return jsonify({
        "logged_in": is_admin(),
        "username":  session.get("admin_user", "") or ((_current_user() or {}).get("email", "")),
    })


# ═══════════════════════════════════════════════════════════════════════════════
# §9B AUTH ROUTES — session-based users (HR/HOD/ADMIN)
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/auth/login", methods=["POST"])
def login():
    d = request.get_json(silent=True) or {}
    email = (d.get("email") or "").strip().lower()
    password = d.get("password") or ""
    if not email or not password:
        return jsonify({"error": "Email and password required."}), 400

    with get_conn() as conn:
        u = get_user_by_email(conn, email)
    if not u or not u.get("is_active"):
        return jsonify({"error": "Invalid credentials."}), 401
    if not check_password_hash(u.get("password_hash", ""), password):
        return jsonify({"error": "Invalid credentials."}), 401

    session.permanent = True
    session["user_id"] = int(u["id"])
    if u.get("role") == "ADMIN":
        session["admin_logged_in"] = True
        session["admin_user"] = u.get("email", "")
    return jsonify({
        "success": True,
        "user": {
            "id": u["id"],
            "name": u["name"],
            "email": u["email"],
            "role": u["role"],
            "department": u.get("department"),
        }
    })


@app.route("/auth/signup", methods=["POST"])
def signup():
    d = request.get_json(silent=True) or {}
    name = (d.get("name") or "").strip()
    email = (d.get("email") or "").strip().lower()
    password = d.get("password") or ""
    role = (d.get("role") or "HR").strip().upper()
    department = (d.get("department") or "").strip() or None

    if not name or not email or not password:
        return jsonify({"error": "Name, email and password required."}), 400
    if len(password) < 8:
        return jsonify({"error": "Password must be at least 8 characters."}), 400
    if role not in ("HR", "HOD", "ADMIN"):
        return jsonify({"error": "Invalid role."}), 400
    if role == "HOD" and not department:
        return jsonify({"error": "Department required for HOD."}), 400
    if role in ("HR", "ADMIN"):
        department = None

    with get_conn() as conn:
        if get_user_by_email(conn, email):
            return jsonify({"error": "Email already registered."}), 409
        uid = insert_user(
            conn,
            name=name,
            email=email,
            password_hash=generate_password_hash(password),
            role=role,
            department=department,
        )

    session.permanent = True
    session["user_id"] = int(uid)
    if role == "ADMIN":
        session["admin_logged_in"] = True
        session["admin_user"] = email
    return jsonify({
        "success": True,
        "user": {
            "id": uid,
            "name": name,
            "email": email,
            "role": role,
            "department": department,
        }
    }), 201


@app.route("/auth/logout", methods=["POST"])
def logout():
    session.clear()
    return jsonify({"success": True})


@app.route("/auth/me", methods=["GET"])
def me():
    u = _current_user()
    if not u:
        return jsonify({"logged_in": False, "user": None})
    return jsonify({
        "logged_in": True,
        "user": {
            "id": u["id"],
            "name": u["name"],
            "email": u["email"],
            "role": u["role"],
            "department": u.get("department"),
        }
    })


# ═══════════════════════════════════════════════════════════════════════════════
# §9C ADMIN USER MANAGEMENT
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/users", methods=["GET"])
@require_roles("ADMIN")
def users_list():
    role = (request.args.get("role") or "").strip().upper() or None
    if role and role not in ("HR", "HOD", "ADMIN"):
        return jsonify({"error": "Invalid role filter."}), 400
    with get_conn() as conn:
        rows = list(db_list_users(conn, role=role))
    out = []
    for r in rows:
        out.append({
            "id": r["id"],
            "name": r["name"],
            "email": r["email"],
            "role": r["role"],
            "department": r.get("department"),
            "is_active": bool(r.get("is_active")),
            "created_at": r.get("created_at"),
        })
    return jsonify(out)


@app.route("/users", methods=["POST"])
@require_roles("ADMIN")
def users_create():
    d = request.get_json(silent=True) or {}
    name = (d.get("name") or "").strip()
    email = (d.get("email") or "").strip().lower()
    password = d.get("password") or ""
    role = (d.get("role") or "").strip().upper()
    department = (d.get("department") or "").strip() or None

    if not name or not email or not password or not role:
        return jsonify({"error": "name, email, password, role required."}), 400
    if role not in ("HR", "HOD", "ADMIN"):
        return jsonify({"error": "Invalid role."}), 400
    if len(password) < 8:
        return jsonify({"error": "Password must be at least 8 characters."}), 400
    if role == "HOD" and not department:
        return jsonify({"error": "Department required for HOD."}), 400
    if role in ("HR", "ADMIN"):
        department = None

    with get_conn() as conn:
        if get_user_by_email(conn, email):
            return jsonify({"error": "Email already registered."}), 409
        uid = insert_user(
            conn,
            name=name,
            email=email,
            password_hash=generate_password_hash(password),
            role=role,
            department=department,
        )
        user = get_user_by_id(conn, uid)

    return jsonify({
        "success": True,
        "user": {
            "id": user["id"],
            "name": user["name"],
            "email": user["email"],
            "role": user["role"],
            "department": user.get("department"),
            "is_active": bool(user.get("is_active")),
            "created_at": user.get("created_at"),
        }
    }), 201


@app.route("/users/<int:user_id>", methods=["DELETE"])
@require_roles("ADMIN")
def users_delete(user_id):
    cur_uid = session.get("user_id")
    try:
        cur_uid = int(cur_uid) if cur_uid is not None else None
    except Exception:
        cur_uid = None
    if cur_uid == user_id:
        return jsonify({"error": "Cannot delete currently logged-in admin."}), 400

    with get_conn() as conn:
        deleted = db_delete_user(conn, user_id)
    if deleted:
        return jsonify({"success": True})
    return jsonify({"error": "Not found"}), 404

@app.route("/admin/change-password", methods=["POST"])
def change_password():
    if not is_admin():
        return jsonify({"error": "Unauthorized"}), 401
    d       = request.get_json(silent=True) or {}
    current = d.get("current_password", "")
    new_pw  = d.get("new_password", "")
    u       = session.get("admin_user", "")
    if not new_pw or len(new_pw) < 8:
        return jsonify({"error": "New password must be at least 8 characters."}), 400
    if ADMIN_STORE.get(u) != _hash(current):
        return jsonify({"error": "Current password is incorrect."}), 401
    ADMIN_STORE[u] = _hash(new_pw)
    return jsonify({"success": True})

# ═══════════════════════════════════════════════════════════════════════════════
# §10 KEYWORD MANAGEMENT
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/admin/keywords", methods=["GET"])
def get_keywords():
    # Backward-compatible admin path; role-based access (HR/ADMIN)
    u = _current_user()
    if not u:
        return jsonify({"error": "Unauthorized"}), 401
    if u.get("role") not in ("HR", "ADMIN"):
        return jsonify({"error": "Forbidden"}), 403
    with get_conn() as conn:
        return jsonify(db_get_all_keywords(conn))

@app.route("/admin/keywords/<department>", methods=["POST"])
def add_keyword(department):
    u = _current_user()
    if not u:
        return jsonify({"error": "Unauthorized"}), 401
    if u.get("role") not in ("HR", "ADMIN"):
        return jsonify({"error": "Forbidden"}), 403

    body = request.get_json(silent=True) or {}
    kw = (body.get("keyword") or "").strip().lower()
    kws = body.get("keywords")
    if kw:
        keywords = [kw]
    elif isinstance(kws, list):
        keywords = kws
    else:
        return jsonify({"error": "keyword or keywords[] required"}), 400

    with get_conn() as conn:
        current = set(db_get_keywords_by_department(conn, department))
        for k in keywords:
            k = (k or "").strip().lower()
            if k:
                current.add(k)
        db_set_keywords_for_department(conn, department, sorted(current))
        updated = db_get_keywords_by_department(conn, department)

    KEYWORD_STORE[department] = updated
    return jsonify({"success": True, "keywords": updated})

@app.route("/admin/keywords/<department>/<keyword>", methods=["DELETE"])
def delete_keyword(department, keyword):
    u = _current_user()
    if not u:
        return jsonify({"error": "Unauthorized"}), 401
    if u.get("role") not in ("HR", "ADMIN"):
        return jsonify({"error": "Forbidden"}), 403

    kw = (keyword or "").strip().lower()
    with get_conn() as conn:
        current = [k for k in db_get_keywords_by_department(conn, department) if k != kw]
        db_set_keywords_for_department(conn, department, current)
        updated = db_get_keywords_by_department(conn, department)

    KEYWORD_STORE[department] = updated
    return jsonify({"success": True, "keywords": updated})


# Public/role-friendly keyword APIs
@app.route("/keywords", methods=["GET"])
@require_roles("HR", "HOD", "ADMIN")
def view_keywords():
    u = _current_user()
    dept_req = (request.args.get("department") or "").strip() or None
    if u.get("role") == "HOD":
        dept_req = u.get("department")
    with get_conn() as conn:
        if dept_req:
            return jsonify({"department": dept_req, "keywords": db_get_keywords_by_department(conn, dept_req)})
        return jsonify(db_get_all_keywords(conn))


@app.route("/keywords/<department>", methods=["POST"])
@require_roles("HR", "ADMIN")
def update_keywords(department):
    body = request.get_json(silent=True) or {}
    kws = body.get("keywords")
    kw = (body.get("keyword") or "").strip().lower()
    if isinstance(kws, list):
        new_list = [str(x).strip().lower() for x in kws if str(x).strip()]
    elif kw:
        new_list = None
    else:
        return jsonify({"error": "keyword or keywords[] required"}), 400

    with get_conn() as conn:
        if new_list is not None:
            db_set_keywords_for_department(conn, department, new_list)
        else:
            current = set(db_get_keywords_by_department(conn, department))
            current.add(kw)
            db_set_keywords_for_department(conn, department, sorted(current))
        updated = db_get_keywords_by_department(conn, department)

    KEYWORD_STORE[department] = updated
    return jsonify({"success": True, "department": department, "keywords": updated})

# ═══════════════════════════════════════════════════════════════════════════════
# §11 DATASET MANAGEMENT
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/admin/dataset", methods=["GET"])
def get_dataset():
    if not is_admin(): return jsonify({"error": "Unauthorized"}), 401
    summary = {}
    for s in CUSTOM_DATASET:
        summary[s["label"]] = summary.get(s["label"], 0) + 1
    return jsonify({
        "custom_total":   len(CUSTOM_DATASET),
        "base_total":     len(BASE_TRAINING_DATA),
        "merged_total":   len(BASE_TRAINING_DATA) + len(CUSTOM_DATASET),
        "by_dept":        summary,
        "model_accuracy": MODEL_ACCURACY,
    })

@app.route("/admin/dataset/upload", methods=["POST"])
def upload_dataset_csv():
    if not is_admin(): return jsonify({"error": "Unauthorized"}), 401
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    f     = request.files["file"]
    added = 0
    try:
        content = f.read().decode("utf-8", errors="ignore")
        reader  = csv.DictReader(io.StringIO(content))
        for row in reader:
            text  = row.get("text",  "").strip()
            label = row.get("label", "").strip()
            if text and label:
                CUSTOM_DATASET.append({"text": text, "label": label})
                added += 1
    except Exception as e:
        return jsonify({"error": str(e)}), 400
    return jsonify({"success": True, "added": added, "total": len(CUSTOM_DATASET)})

@app.route("/admin/dataset/add", methods=["POST"])
def add_dataset_samples():
    if not is_admin(): return jsonify({"error": "Unauthorized"}), 401
    samples = (request.get_json(silent=True) or {}).get("samples", [])
    added   = sum(1 for s in samples
                  if s.get("text") and s.get("label")
                  and not CUSTOM_DATASET.append({"text": s["text"], "label": s["label"]}))
    return jsonify({"success": True, "added": added, "total": len(CUSTOM_DATASET)})

@app.route("/admin/dataset/retrain", methods=["POST"])
def retrain_model():
    if not is_admin(): return jsonify({"error": "Unauthorized"}), 401
    global model, vectorizer, MODEL_ACCURACY
    extra_texts  = [s["text"]  for s in CUSTOM_DATASET]
    extra_labels = [s["label"] for s in CUSTOM_DATASET]
    try:
        model, vectorizer, MODEL_ACCURACY = build_model(extra_texts, extra_labels)
        pickle.dump(model,      open("model.pkl",      "wb"))
        pickle.dump(vectorizer, open("vectorizer.pkl", "wb"))
        return jsonify({
            "success":       True,
            "accuracy":      MODEL_ACCURACY,
            "total_samples": len(BASE_TRAINING_DATA) + len(CUSTOM_DATASET),
            "custom_added":  len(CUSTOM_DATASET),
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/admin/dataset/clear", methods=["DELETE"])
def clear_custom_dataset():
    if not is_admin(): return jsonify({"error": "Unauthorized"}), 401
    CUSTOM_DATASET.clear()
    return jsonify({"success": True})

# ═══════════════════════════════════════════════════════════════════════════════
# §12 SINGLE PREDICT
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/predict", methods=["POST"])
def predict():
    if "resume" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    file      = request.files["resume"]
    fname     = file.filename.lower()
    raw_bytes = file.read()

    if fname.endswith(".pdf"):
        try: text = extract_pdf(io.BytesIO(raw_bytes))
        except Exception as e: return jsonify({"error": str(e)}), 400
    elif fname.endswith(".docx"):
        try: text = extract_docx(io.BytesIO(raw_bytes))
        except Exception as e: return jsonify({"error": str(e)}), 400
    else:
        return jsonify({"error": "Upload PDF or DOCX only."}), 400

    valid, reason = is_resume(text)
    if not valid:
        return jsonify({"error": reason, "not_a_resume": True}), 422

    clf = pickle.load(open("model.pkl",      "rb"))
    vec = pickle.load(open("vectorizer.pkl", "rb"))
    r   = full_parse(text, file.filename, raw_bytes, clf, vec)

    # Persist to DB (source of truth)
    uploaded_by = session.get("user_id")
    try:
        uploaded_by = int(uploaded_by) if uploaded_by is not None else None
    except Exception:
        uploaded_by = None

    cloud = CLOUD_STORE.get(r["id"]) or {}
    with get_conn() as conn:
        db_insert_resume(conn, {
            "id": r["id"],
            "original_filename": file.filename,
            "stored_path": cloud.get("stored_path") or "",
            "mime": file.mimetype,
            "size": len(raw_bytes),
            "uploaded_by": uploaded_by,
            "candidate_name": r.get("name"),
            "predicted_department": r.get("department"),
            "probabilities": r.get("probabilities") or {},
            "confidence": r.get("confidence") or 0,
            "kw_score": r.get("kw_score") or 0,
            "relevance": r.get("relevance") or 0,
            "shortlisted": bool(r.get("shortlisted")),
            "skills": r.get("skills") or [],
            "education": r.get("education") or [],
            "experience": r.get("experience") or "Not specified",
            "created_at": datetime.utcnow().replace(microsecond=0).isoformat() + "Z",
        })

    if r["shortlisted"]:
        _store_shortlisted(r)
    return jsonify(r)

# ═══════════════════════════════════════════════════════════════════════════════
# §13 BULK PREDICT — handles individual files, ZIP, and folder (multi-select)
# ═══════════════════════════════════════════════════════════════════════════════

def _process_file_bytes(fname_orig, raw_bytes, clf, vec):
    """Returns (result_dict, error_string). Exactly one will be non-None."""
    fname = fname_orig.lower()
    try:
        if fname.endswith(".pdf"):
            text = extract_pdf(io.BytesIO(raw_bytes))
        elif fname.endswith(".docx"):
            text = extract_docx(io.BytesIO(raw_bytes))
        else:
            return None, f"Unsupported format: {fname_orig}"
    except Exception as e:
        return None, str(e)

    valid, reason = is_resume(text)
    if not valid:
        return None, reason

    r = full_parse(text, fname_orig, raw_bytes, clf, vec)
    return r, None


@app.route("/predict/bulk", methods=["POST"])
@require_login
def predict_bulk():
    files = request.files.getlist("resumes")
    if not files:
        return jsonify({"error": "No files uploaded"}), 400

    clf = pickle.load(open("model.pkl",      "rb"))
    vec = pickle.load(open("vectorizer.pkl", "rb"))

    results, errors = [], []
    total_submitted = 0

    for f in files:
        raw_bytes  = f.read()
        fname_low  = f.filename.lower()

        # — ZIP file handling —
        if fname_low.endswith(".zip"):
            try:
                with zipfile.ZipFile(io.BytesIO(raw_bytes)) as zf:
                    for zname in zf.namelist():
                        # Skip directories and hidden files
                        if zname.endswith("/") or os.path.basename(zname).startswith("."):
                            continue
                        zname_low = zname.lower()
                        if not (zname_low.endswith(".pdf") or zname_low.endswith(".docx")):
                            continue
                        total_submitted += 1
                        if total_submitted > 1000:
                            errors.append({"file": zname, "error": "Batch limit 1000 exceeded"})
                            continue
                        try:
                            zbytes = zf.read(zname)
                            r, err = _process_file_bytes(os.path.basename(zname), zbytes, clf, vec)
                            if err:
                                errors.append({"file": zname, "error": err})
                            else:
                                results.append(r)
                                cloud = CLOUD_STORE.get(r["id"]) or {}
                                uploaded_by = session.get("user_id")
                                try:
                                    uploaded_by = int(uploaded_by) if uploaded_by is not None else None
                                except Exception:
                                    uploaded_by = None
                                mime = "application/pdf" if zname_low.endswith(".pdf") else "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
                                with get_conn() as conn:
                                    db_insert_resume(conn, {
                                        "id": r["id"],
                                        "original_filename": os.path.basename(zname),
                                        "stored_path": cloud.get("stored_path") or "",
                                        "mime": mime,
                                        "size": len(zbytes),
                                        "uploaded_by": uploaded_by,
                                        "candidate_name": r.get("name"),
                                        "predicted_department": r.get("department"),
                                        "probabilities": r.get("probabilities") or {},
                                        "confidence": r.get("confidence") or 0,
                                        "kw_score": r.get("kw_score") or 0,
                                        "relevance": r.get("relevance") or 0,
                                        "shortlisted": bool(r.get("shortlisted")),
                                        "skills": r.get("skills") or [],
                                        "education": r.get("education") or [],
                                        "experience": r.get("experience") or "Not specified",
                                        "created_at": datetime.utcnow().replace(microsecond=0).isoformat() + "Z",
                                    })
                                if r["shortlisted"]:
                                    _store_shortlisted(r)
                        except Exception as e:
                            errors.append({"file": zname, "error": str(e)})
            except zipfile.BadZipFile:
                errors.append({"file": f.filename, "error": "Invalid ZIP file"})
            continue

        # — Regular PDF / DOCX —
        total_submitted += 1
        if total_submitted > 1000:
            errors.append({"file": f.filename, "error": "Batch limit 1000 exceeded"})
            continue
        r, err = _process_file_bytes(f.filename, raw_bytes, clf, vec)
        if err:
            errors.append({"file": f.filename, "error": err})
        else:
            results.append(r)
            cloud = CLOUD_STORE.get(r["id"]) or {}
            uploaded_by = session.get("user_id")
            try:
                uploaded_by = int(uploaded_by) if uploaded_by is not None else None
            except Exception:
                uploaded_by = None
            with get_conn() as conn:
                db_insert_resume(conn, {
                    "id": r["id"],
                    "original_filename": f.filename,
                    "stored_path": cloud.get("stored_path") or "",
                    "mime": f.mimetype,
                    "size": len(raw_bytes),
                    "uploaded_by": uploaded_by,
                    "candidate_name": r.get("name"),
                    "predicted_department": r.get("department"),
                    "probabilities": r.get("probabilities") or {},
                    "confidence": r.get("confidence") or 0,
                    "kw_score": r.get("kw_score") or 0,
                    "relevance": r.get("relevance") or 0,
                    "shortlisted": bool(r.get("shortlisted")),
                    "skills": r.get("skills") or [],
                    "education": r.get("education") or [],
                    "experience": r.get("experience") or "Not specified",
                    "created_at": datetime.utcnow().replace(microsecond=0).isoformat() + "Z",
                })
            if r["shortlisted"]:
                _store_shortlisted(r)

    return jsonify({
        "total":       total_submitted,
        "processed":   len(results),
        "shortlisted": sum(1 for r in results if r["shortlisted"]),
        "errors":      len(errors),
        "results":     results,
        "error_list":  errors,
    })


# ═══════════════════════════════════════════════════════════════════════════════
# §13B RESUME LIST / DOWNLOAD / DELETE (DB-backed)
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/resumes", methods=["GET"])
@require_roles("HR", "HOD", "ADMIN")
def list_resumes():
    dept = request.args.get("department")
    shortlisted_q = request.args.get("shortlisted")
    search = request.args.get("search")
    sort = request.args.get("sort", "created_at")

    try:
        page = int(request.args.get("page", 1))
    except Exception:
        page = 1
    try:
        page_size = int(request.args.get("page_size", 25))
    except Exception:
        page_size = 25
    page = max(1, page)
    page_size = min(max(1, page_size), 200)

    shortlisted = None
    if isinstance(shortlisted_q, str) and shortlisted_q != "":
        if shortlisted_q.lower() in ("true", "1", "yes"):
            shortlisted = True
        elif shortlisted_q.lower() in ("false", "0", "no"):
            shortlisted = False

    requested_dept = (dept or "").strip() or None
    effective_dept, forced_dept = _effective_department_filter(requested_dept)

    with get_conn() as conn:
        total, rows = db_query_resumes(
            conn,
            department=effective_dept,
            shortlisted=shortlisted,
            search=search,
            sort=sort,
            page=page,
            page_size=page_size,
            forced_department=forced_dept,
        )

    return jsonify({
        "total": total,
        "page": page,
        "page_size": page_size,
        "results": [_resume_row_to_payload(r) for r in rows],
    })


@app.route("/resumes/<resume_id>/download", methods=["GET"])
@require_roles("HR", "HOD", "ADMIN")
def download_resume(resume_id):
    u = _current_user()
    with get_conn() as conn:
        row = db_get_resume_by_id(conn, resume_id)
    if not row:
        return jsonify({"error": "Not found"}), 404
    if not _can_user_access_resume(u, row):
        return jsonify({"error": "Forbidden"}), 403

    path = row.get("stored_path")
    if not path or not os.path.exists(path):
        return jsonify({"error": "File missing"}), 410

    return send_file(
        path,
        as_attachment=True,
        download_name=row.get("original_filename") or os.path.basename(path),
    )


@app.route("/resumes/<resume_id>", methods=["DELETE"])
@require_roles("HR", "ADMIN")
def delete_resume(resume_id):
    with get_conn() as conn:
        row = db_get_resume_by_id(conn, resume_id)
        if not row:
            return jsonify({"error": "Not found"}), 404
        deleted = db_delete_resume(conn, resume_id)
    if deleted:
        path = row.get("stored_path")
        if path and os.path.exists(path):
            try:
                os.remove(path)
            except Exception:
                pass
        return jsonify({"success": True})
    return jsonify({"error": "Not found"}), 404

# ═══════════════════════════════════════════════════════════════════════════════
# §14 SHORTLISTED ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/admin/shortlisted", methods=["GET"])
def get_shortlisted():
    u = _current_user()
    if not u:
        return jsonify({"error": "Unauthorized"}), 401
    if u.get("role") not in ("HR", "HOD", "ADMIN"):
        return jsonify({"error": "Forbidden"}), 403

    dept_req = (request.args.get("department") or "").strip() or None
    sort_by = request.args.get("sort", "relevance")
    dept_effective, dept_forced = _effective_department_filter(dept_req)

    with get_conn() as conn:
        _total, rows = db_query_resumes(
            conn,
            department=dept_effective,
            shortlisted=True,
            search=None,
            sort=sort_by,
            page=1,
            page_size=10000,
            forced_department=dept_forced,
        )
    data = [_resume_row_to_payload(r) for r in rows]

    key_map = {"relevance": "relevance", "confidence": "confidence", "kw_score": "kw_score"}
    data.sort(key=lambda x: x.get(key_map.get(sort_by, "relevance"), 0) or 0, reverse=True)

    dept_rank = {}
    out = []
    for c in data:
        d = c["department"]
        dept_rank[d] = dept_rank.get(d, 0) + 1
        row = dict(c)
        row["dept_rank"] = dept_rank[d]
        out.append(row)
    return jsonify(out)

@app.route("/admin/shortlisted/export/csv", methods=["GET"])
def export_csv():
    u = _current_user()
    if not u:
        return jsonify({"error": "Unauthorized"}), 401
    if u.get("role") not in ("HR", "HOD", "ADMIN"):
        return jsonify({"error": "Forbidden"}), 403

    dept_effective, dept_forced = _effective_department_filter((request.args.get("department") or "").strip() or None)
    with get_conn() as conn:
        _total, rows = db_query_resumes(
            conn,
            department=dept_effective,
            shortlisted=True,
            search=None,
            sort="relevance",
            page=1,
            page_size=10000,
            forced_department=dept_forced,
        )
    data = [_resume_row_to_payload(r) for r in rows]
    data = sorted(data, key=lambda x: (x["department"], -(x.get("relevance") or 0)))
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Rank","Candidate Name","Department","Dept Rank",
        "Relevance Score (%)","ML Confidence (%)","Keyword Score (%)",
        "Top Skills","Education","Experience",
        "Resume Link","Resume ID"
    ])
    dept_rank_tracker = {}
    base = request.host_url.rstrip("/")
    for i, c in enumerate(data, 1):
        d = c["department"]
        dept_rank_tracker[d] = dept_rank_tracker.get(d, 0) + 1
        resume_link = base + f"/resumes/{c.get('id','')}/download" if c.get("id") else ""
        writer.writerow([
            i, c.get("name",""), d, dept_rank_tracker[d],
            c.get("relevance",""), c.get("confidence",""), c.get("kw_score",""),
            " | ".join(c.get("skills", [])),
            " | ".join(c.get("education", [])),
            c.get("experience",""),
            resume_link,
            c.get("id",""),
        ])
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment;filename=shortlisted_candidates.csv"}
    )

@app.route("/admin/shortlisted/export/excel", methods=["GET"])
def export_excel():
    """Export shortlisted candidates to Excel with clickable hyperlinks."""
    u = _current_user()
    if not u:
        return jsonify({"error": "Unauthorized"}), 401
    if u.get("role") not in ("HR", "HOD", "ADMIN"):
        return jsonify({"error": "Forbidden"}), 403
    if not EXCEL_OK:
        return jsonify({"error": "openpyxl not installed. Run: pip install openpyxl"}), 500

    dept_effective, dept_forced = _effective_department_filter((request.args.get("department") or "").strip() or None)
    with get_conn() as conn:
        _total, rows = db_query_resumes(
            conn,
            department=dept_effective,
            shortlisted=True,
            search=None,
            sort="relevance",
            page=1,
            page_size=10000,
            forced_department=dept_forced,
        )
    data = [_resume_row_to_payload(r) for r in rows]
    data = sorted(data, key=lambda x: (x["department"], -(x.get("relevance") or 0)))
    wb   = openpyxl.Workbook()
    ws   = wb.active
    ws.title = "Shortlisted Candidates"

    # ── Styles ────────────────────────────────────────────────────────────────
    header_fill  = PatternFill("solid", fgColor="0A1628")
    header_font  = Font(bold=True, color="FFFFFF", size=10)
    link_font    = Font(color="1D4ED8", underline="single", size=10)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin_border  = Border(
        left=Side(style="thin", color="E2E8F0"), right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"), bottom=Side(style="thin", color="E2E8F0"),
    )

    headers = [
        "Rank", "Candidate Name", "Department", "Dept Rank",
        "Relevance (%)", "ML Conf (%)", "KW Score (%)",
        "Top Skills", "Education", "Experience",
        "Resume Link", "Resume ID", "Timestamp"
    ]
    col_widths = [6, 22, 14, 10, 12, 12, 12, 35, 30, 22, 50, 12, 17]

    ws.row_dimensions[1].height = 22
    for col_idx, (hdr, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=hdr)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center_align
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    dept_colors = {
        "Data Science": "F0FDF4",
        "Developer":    "EFF6FF",
        "HR":           "FDF4FF",
        "Finance":      "FFF7ED",
        "Testing":      "FEFCE8",
    }
    dept_rank_tracker = {}
    base = request.host_url.rstrip("/")

    for row_i, c in enumerate(data, 2):
        d  = c["department"]
        dept_rank_tracker[d] = dept_rank_tracker.get(d, 0) + 1
        row_fill    = PatternFill("solid", fgColor=dept_colors.get(d, "FFFFFF"))
        row_values  = [
            row_i - 1,
            c.get("name",         ""),
            d,
            dept_rank_tracker[d],
            c.get("relevance",    ""),
            c.get("confidence",   ""),
            c.get("kw_score",     ""),
            " | ".join(c.get("skills",    [])),
            " | ".join(c.get("education", [])),
            c.get("experience",   ""),
            None,                               # resume link — handled separately
            c.get("id",           ""),
            c.get("timestamp",    ""),
        ]
        ws.row_dimensions[row_i].height = 18
        for col_i, val in enumerate(row_values, 1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.fill      = row_fill
            cell.border    = thin_border
            cell.alignment = left_align if col_i > 1 else center_align

        # ── Clickable hyperlink in Resume Link column (col 11) ────────────────
        link = base + f"/resumes/{c.get('id','')}/download" if c.get("id") else ""
        if link:
            link_cell                  = ws.cell(row=row_i, column=11, value="📎 Open Resume")
            link_cell.hyperlink        = link
            link_cell.font             = link_font
            link_cell.fill             = row_fill
            link_cell.border           = thin_border
            link_cell.alignment        = center_align

    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="shortlisted_candidates.xlsx",
    )

@app.route("/admin/shortlisted/export/json", methods=["GET"])
def export_json():
    u = _current_user()
    if not u:
        return jsonify({"error": "Unauthorized"}), 401
    if u.get("role") not in ("HR", "HOD", "ADMIN"):
        return jsonify({"error": "Forbidden"}), 403
    grouped = {}
    dept_effective, dept_forced = _effective_department_filter((request.args.get("department") or "").strip() or None)
    with get_conn() as conn:
        _total, rows = db_query_resumes(
            conn,
            department=dept_effective,
            shortlisted=True,
            search=None,
            sort="relevance",
            page=1,
            page_size=10000,
            forced_department=dept_forced,
        )
    base = request.host_url.rstrip("/")
    payloads = [_resume_row_to_payload(r) for r in rows]
    for c in sorted(payloads, key=lambda x: (x["department"], -(x.get("relevance") or 0))):
        c = dict(c)
        if c.get("id"):
            c["resume_link"] = base + f"/resumes/{c['id']}/download"
        grouped.setdefault(c["department"], []).append(c)
    return jsonify({"exported_at": datetime.now().isoformat(), "departments": grouped})


# Non-admin-friendly export aliases (same behavior/scoping)
@app.route("/resumes/export/csv", methods=["GET"])
def export_resumes_csv_alias():
    u = _current_user()
    if not u:
        return jsonify({"error": "Unauthorized"}), 401
    if u.get("role") not in ("HR", "HOD", "ADMIN"):
        return jsonify({"error": "Forbidden"}), 403

    dept_effective, dept_forced = _effective_department_filter((request.args.get("department") or "").strip() or None)
    shortlisted_q = request.args.get("shortlisted")
    search = request.args.get("search")
    sort = request.args.get("sort", "created_at")

    shortlisted = None
    if isinstance(shortlisted_q, str) and shortlisted_q != "":
        if shortlisted_q.lower() in ("true", "1", "yes"):
            shortlisted = True
        elif shortlisted_q.lower() in ("false", "0", "no"):
            shortlisted = False

    with get_conn() as conn:
        _total, rows = db_query_resumes(
            conn,
            department=dept_effective,
            shortlisted=shortlisted,
            search=search,
            sort=sort,
            page=1,
            page_size=100000,
            forced_department=dept_forced,
        )

    data = [_resume_row_to_payload(r) for r in rows]
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Candidate Name","Department",
        "Relevance Score (%)","ML Confidence (%)","Keyword Score (%)",
        "Top Skills","Education","Experience",
        "Shortlisted","Resume Link","Resume ID","File"
    ])
    base = request.host_url.rstrip("/")
    for c in data:
        resume_link = base + f"/resumes/{c.get('id','')}/download" if c.get("id") else ""
        writer.writerow([
            c.get("name",""),
            c.get("department",""),
            c.get("relevance",""),
            c.get("confidence",""),
            c.get("kw_score",""),
            " | ".join(c.get("skills", [])),
            " | ".join(c.get("education", [])),
            c.get("experience",""),
            "Yes" if c.get("shortlisted") else "No",
            resume_link,
            c.get("id",""),
            c.get("file",""),
        ])
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment;filename=resumes_export.csv"}
    )


@app.route("/resumes/export/excel", methods=["GET"])
def export_resumes_excel_alias():
    u = _current_user()
    if not u:
        return jsonify({"error": "Unauthorized"}), 401
    if u.get("role") not in ("HR", "HOD", "ADMIN"):
        return jsonify({"error": "Forbidden"}), 403
    if not EXCEL_OK:
        return jsonify({"error": "openpyxl not installed. Run: pip install openpyxl"}), 500

    dept_effective, dept_forced = _effective_department_filter((request.args.get("department") or "").strip() or None)
    shortlisted_q = request.args.get("shortlisted")
    search = request.args.get("search")
    sort = request.args.get("sort", "created_at")

    shortlisted = None
    if isinstance(shortlisted_q, str) and shortlisted_q != "":
        if shortlisted_q.lower() in ("true", "1", "yes"):
            shortlisted = True
        elif shortlisted_q.lower() in ("false", "0", "no"):
            shortlisted = False

    with get_conn() as conn:
        _total, rows = db_query_resumes(
            conn,
            department=dept_effective,
            shortlisted=shortlisted,
            search=search,
            sort=sort,
            page=1,
            page_size=100000,
            forced_department=dept_forced,
        )
    data = [_resume_row_to_payload(r) for r in rows]

    wb   = openpyxl.Workbook()
    ws   = wb.active
    ws.title = "Resumes"

    header_fill  = PatternFill("solid", fgColor="0A1628")
    header_font  = Font(bold=True, color="FFFFFF", size=10)
    link_font    = Font(color="1D4ED8", underline="single", size=10)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin_border  = Border(
        left=Side(style="thin", color="E2E8F0"), right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"), bottom=Side(style="thin", color="E2E8F0"),
    )

    headers = [
        "Candidate Name","Department",
        "Relevance (%)","ML Conf (%)","KW Score (%)",
        "Top Skills","Education","Experience",
        "Shortlisted","Resume Link","Resume ID","Timestamp","File"
    ]
    ws.row_dimensions[1].height = 22
    for col_idx, hdr in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=hdr)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center_align
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = 20 if col_idx < 6 else 35

    base = request.host_url.rstrip("/")
    for row_i, c in enumerate(data, 2):
        values = [
            c.get("name",""),
            c.get("department",""),
            c.get("relevance",""),
            c.get("confidence",""),
            c.get("kw_score",""),
            " | ".join(c.get("skills", [])),
            " | ".join(c.get("education", [])),
            c.get("experience",""),
            "Yes" if c.get("shortlisted") else "No",
            None,
            c.get("id",""),
            c.get("timestamp",""),
            c.get("file",""),
        ]
        for col_i, val in enumerate(values, 1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.border = thin_border
            cell.alignment = left_align if col_i not in (3,4,5,9) else center_align

        link = base + f"/resumes/{c.get('id','')}/download" if c.get("id") else ""
        if link:
            link_cell = ws.cell(row=row_i, column=10, value="⬇️ Download")
            link_cell.hyperlink = link
            link_cell.font = link_font
            link_cell.border = thin_border
            link_cell.alignment = center_align

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="resumes_export.xlsx",
    )


@app.route("/resumes/export/json", methods=["GET"])
def export_resumes_json_alias():
    u = _current_user()
    if not u:
        return jsonify({"error": "Unauthorized"}), 401
    if u.get("role") not in ("HR", "HOD", "ADMIN"):
        return jsonify({"error": "Forbidden"}), 403

    dept_effective, dept_forced = _effective_department_filter((request.args.get("department") or "").strip() or None)
    shortlisted_q = request.args.get("shortlisted")
    search = request.args.get("search")
    sort = request.args.get("sort", "created_at")

    shortlisted = None
    if isinstance(shortlisted_q, str) and shortlisted_q != "":
        if shortlisted_q.lower() in ("true", "1", "yes"):
            shortlisted = True
        elif shortlisted_q.lower() in ("false", "0", "no"):
            shortlisted = False

    with get_conn() as conn:
        _total, rows = db_query_resumes(
            conn,
            department=dept_effective,
            shortlisted=shortlisted,
            search=search,
            sort=sort,
            page=1,
            page_size=100000,
            forced_department=dept_forced,
        )
    base = request.host_url.rstrip("/")
    payloads = [_resume_row_to_payload(r) for r in rows]
    for c in payloads:
        if c.get("id"):
            c["resume_link"] = base + f"/resumes/{c['id']}/download"
    return jsonify({"exported_at": datetime.now().isoformat(), "total": len(payloads), "results": payloads})

@app.route("/admin/shortlisted/clear", methods=["DELETE"])
def clear_shortlisted():
    if not is_admin(): return jsonify({"error": "Unauthorized"}), 401
    SHORTLISTED.clear()
    return jsonify({"success": True})

# ═══════════════════════════════════════════════════════════════════════════════
# §15 CLOUD STORAGE ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/admin/cloud", methods=["GET"])
def list_cloud():
    if not is_admin(): return jsonify({"error": "Unauthorized"}), 401
    return jsonify({"total": len(CLOUD_STORE), "files": list(CLOUD_STORE.values())})

@app.route("/admin/cloud/<resume_id>", methods=["GET"])
def get_cloud_link(resume_id):
    if not is_admin(): return jsonify({"error": "Unauthorized"}), 401
    entry = CLOUD_STORE.get(resume_id)
    if not entry: return jsonify({"error": "Not found"}), 404
    return jsonify(entry)

# ═══════════════════════════════════════════════════════════════════════════════
# §16 STATS + THRESHOLD
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/admin/stats")
def stats():
    if not is_admin(): return jsonify({"error": "Unauthorized"}), 401
    dept_counts, dept_avg = {}, {}
    for c in SHORTLISTED:
        d = c["department"]
        dept_counts[d] = dept_counts.get(d, 0) + 1
        dept_avg[d]    = dept_avg.get(d, []) + [c.get("relevance", 0)]
    return jsonify({
        "total_shortlisted":     len(SHORTLISTED),
        "by_department":         dept_counts,
        "avg_relevance_by_dept": {d: round(float(np.mean(v)), 1) for d, v in dept_avg.items()},
        "avg_confidence":  round(float(np.mean([c["confidence"]      for c in SHORTLISTED])), 1) if SHORTLISTED else 0,
        "avg_kw_score":    round(float(np.mean([c["kw_score"]        for c in SHORTLISTED])), 1) if SHORTLISTED else 0,
        "avg_relevance":   round(float(np.mean([c.get("relevance",0) for c in SHORTLISTED])), 1) if SHORTLISTED else 0,
        "model_accuracy":        MODEL_ACCURACY,
        "total_cloud_files":     len(CLOUD_STORE),
        "custom_dataset_size":   len(CUSTOM_DATASET),
        "shortlist_threshold":   SHORTLIST_THRESHOLD,
    })


@app.route("/")
def home():
    return send_file(os.path.join(app.root_path, "INDEX.html"))

if __name__ == "__main__":
    app.run(debug=True, port=5000)
